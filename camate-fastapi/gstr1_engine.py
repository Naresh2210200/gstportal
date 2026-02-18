"""
GSTR1 Excel Generation Engine
==============================
Core logic for generating the GSTR1 Excel workbook from CSV data.
This mirrors the frontend gstr1Processor.ts logic but runs server-side
with access to the full Excel template from R2.

Returns: (excel_bytes: bytes, sheets_processed: int)
"""
import io
import csv
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

# ─── Standard GSTR1 column mappings (mirrors frontend gstr1Processor.ts) ─────
COLUMN_MAPPINGS: Dict[str, Dict[str, str]] = {
    "b2b": {
        "GSTIN/UIN of Recipient": "GSTIN/UIN of Recipient",
        "Receiver Name": "Receiver Name",
        "Invoice Number": "Invoice Number",
        "Invoice date": "Invoice date",
        "Invoice Value": "Invoice Value",
        "Place Of Supply": "Place Of Supply",
        "Reverse Charge": "Reverse Charge",
        "Applicable % of Tax Rate": "Applicable % of Tax Rate",
        "Invoice Type": "Invoice Type",
        "E-Commerce GSTIN": "E-Commerce GSTIN",
        "Rate": "Rate",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Central Tax Amount": "Central Tax Amount",
        "State/UT Tax Amount": "State/UT Tax Amount",
        "Cess Amount": "Cess Amount",
    },
    "b2cl": {
        "Invoice Number": "Invoice Number",
        "Invoice date": "Invoice date",
        "Invoice Value": "Invoice Value",
        "Place Of Supply": "Place Of Supply",
        "Applicable % of Tax Rate": "Applicable % of Tax Rate",
        "Rate": "Rate",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Cess Amount": "Cess Amount",
        "E-Commerce GSTIN": "E-Commerce GSTIN",
    },
    "b2cs": {
        "Type": "Type",
        "Place Of Supply": "Place Of Supply",
        "Applicable % of Tax Rate": "Applicable % of Tax Rate",
        "Rate": "Rate",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Central Tax Amount": "Central Tax Amount",
        "State/UT Tax Amount": "State/UT Tax Amount",
        "Cess Amount": "Cess Amount",
        "E-Commerce GSTIN": "E-Commerce GSTIN",
    },
    "hsn": {
        "HSN": "HSN",
        "Description": "Description",
        "UQC": "UQC",
        "Total Quantity": "Total Quantity",
        "Total Value": "Total Value",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Central Tax Amount": "Central Tax Amount",
        "State/UT Tax Amount": "State/UT Tax Amount",
        "Cess Amount": "Cess Amount",
    },
    "Nil_exempt_NonGST": {
        "Description": "Description",
        "Nil Rated Supplies": "Nil Rated Supplies",
        "Exempted (other than nil rated/non GST supply)": "Exempted (other than nil rated/non GST supply)",
        "Non-GST supplies": "Non-GST Supplies",
    },
    "cdnr": {
        "GSTIN/UIN of Recipient": "GSTIN/UIN of Recipient",
        "Receiver Name": "Receiver Name",
        "Note Number": "Note Number",
        "Note date": "Note date",
        "Note Type": "Note Type",
        "Place Of Supply": "Place Of Supply",
        "Reverse Charge": "Reverse Charge",
        "Note Supply Type": "Note Supply Type",
        "Note Value": "Note Value",
        "Applicable % of Tax Rate": "Applicable % of Tax Rate",
        "Rate": "Rate",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Central Tax Amount": "Central Tax Amount",
        "State/UT Tax Amount": "State/UT Tax Amount",
        "Cess Amount": "Cess Amount",
    },
    "cdnur": {
        "UR Type": "UR Type",
        "Note Number": "Note Number",
        "Note date": "Note date",
        "Note Type": "Note Type",
        "Place Of Supply": "Place Of Supply",
        "Note Value": "Note Value",
        "Applicable % of Tax Rate": "Applicable % of Tax Rate",
        "Rate": "Rate",
        "Taxable Value": "Taxable Value",
        "Integrated Tax Amount": "Integrated Tax Amount",
        "Cess Amount": "Cess Amount",
    },
    "Docs_issued": {
        "Nature of Document": "Nature of Document",
        "Sr.No.From": "Sr.No.From",
        "Sr.No.To": "Sr.No.To",
        "Total Number": "Total Number",
        "Cancelled": "Cancelled",
    },
}

# Map CSV file name patterns to GSTR1 sheet names
SHEET_DETECTION: Dict[str, str] = {
    "b2b": "b2b",
    "b2cl": "b2cl",
    "b2cs": "b2cs",
    "hsn": "hsn",
    "nil": "Nil_exempt_NonGST",
    "exempt": "Nil_exempt_NonGST",
    "cdnr": "cdnr",
    "cdnur": "cdnur",
    "docs": "Docs_issued",
    "document": "Docs_issued",
}


def detect_sheet(file_name: str) -> str | None:
    """Detect which GSTR1 sheet a CSV file belongs to based on its name."""
    name_lower = file_name.lower()
    for keyword, sheet in SHEET_DETECTION.items():
        if keyword in name_lower:
            return sheet
    return None


def parse_csv(content: str) -> tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV content into headers and rows."""
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = list(reader)
    return list(headers), rows


def map_row(row: Dict[str, str], sheet_name: str) -> Dict[str, str]:
    """Map a CSV row's keys to standard GSTR1 column names."""
    mapping = COLUMN_MAPPINGS.get(sheet_name, {})
    mapped = {}
    for src_col, target_col in mapping.items():
        # Try exact match first, then case-insensitive
        if src_col in row:
            mapped[target_col] = row[src_col]
        else:
            for k, v in row.items():
                if k.strip().lower() == src_col.lower():
                    mapped[target_col] = v
                    break
    return mapped


def generate_gstr1_excel(csv_files: List[Dict[str, str]]) -> tuple[bytes, int]:
    """
    Generate a GSTR1 Excel workbook from a list of CSV files.
    
    Args:
        csv_files: List of {"name": filename, "content": csv_string}
    
    Returns:
        (excel_bytes, sheets_processed_count)
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    sheets_processed = 0

    for csv_file in csv_files:
        file_name = csv_file["name"]
        content = csv_file["content"]

        sheet_name = detect_sheet(file_name)
        if not sheet_name:
            logger.warning(f"Could not detect sheet for file: {file_name}, skipping.")
            continue

        try:
            headers, rows = parse_csv(content)
        except Exception as e:
            logger.error(f"Failed to parse {file_name}: {e}")
            continue

        if not rows:
            logger.info(f"No data rows in {file_name}, skipping.")
            continue

        # Create or get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            start_row = 1

        # Get target columns for this sheet
        target_columns = list(COLUMN_MAPPINGS.get(sheet_name, {}).values())

        # Write header row (only for new sheets)
        if start_row == 1:
            for col_idx, col_name in enumerate(target_columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            ws.row_dimensions[1].height = 30
            start_row = 2

        # Write data rows
        for row_data in rows:
            mapped = map_row(row_data, sheet_name)
            for col_idx, col_name in enumerate(target_columns, start=1):
                value = mapped.get(col_name, '')
                cell = ws.cell(row=start_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            start_row += 1

        # Auto-fit column widths
        for col_idx, col_name in enumerate(target_columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        sheets_processed += 1
        logger.info(f"Processed sheet '{sheet_name}' from '{file_name}' ({len(rows)} rows)")

    if sheets_processed == 0:
        raise ValueError("No valid GSTR1 sheets could be generated from the provided files.")

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), sheets_processed
